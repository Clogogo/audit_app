"""
School Financial Statements — static Excel parser + dynamic transaction-based computation.
"""
import io
import time
from datetime import date
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from database import get_db
from models import Transaction, Asset, StaffLoan
from routers.assets import calc_accumulated, calc_period_dep
from routers.staff_loans import _outstanding as _staff_loan_outstanding

router = APIRouter(prefix="/financial-statements", tags=["financial-statements"])

# ── Simple TTL cache for computed statements (avoids repeat DB work) ───────────
_RESULT_CACHE: dict[str, tuple[float, dict]] = {}
_CACHE_TTL = 30  # seconds — invalidated naturally when server restarts


def _cache_get(key: str) -> dict | None:
    entry = _RESULT_CACHE.get(key)
    if entry and time.monotonic() - entry[0] < _CACHE_TTL:
        return entry[1]
    return None


def _cache_set(key: str, value: dict) -> None:
    _RESULT_CACHE[key] = (time.monotonic(), value)


def _cache_bust() -> None:
    _RESULT_CACHE.clear()

_FS_FILE = Path(__file__).parent.parent.parent.parent / "MLS_FinancialStatements_FINAL.xlsx"


def _rows(ws, max_row: int | None = None) -> list[tuple]:
    """Return non-empty rows from a worksheet."""
    kwargs: dict[str, Any] = {"values_only": True}
    if max_row:
        kwargs["max_row"] = max_row
    return [r for r in ws.iter_rows(**kwargs) if any(v is not None for v in r)]


def _parse_pl(ws) -> dict:
    rows = _rows(ws)
    result: dict[str, Any] = {
        "company": rows[0][0],
        "title": rows[1][0],
        "period": rows[2][0],
        "income_items": [],
        "total_income": None,
        "expense_items": [],
        "total_expenses": None,
        "operating_profit": None,
        "finance_cost_items": [],
        "total_finance_costs": None,
        "profit_before_tax": None,
        "tax_items": [],
        "total_tax": None,
        "net_profit": None,
        "notes": [],
    }

    section: str | None = None
    for row in rows[4:]:
        first = str(row[0]).strip() if row[0] else ""
        line_amount = row[2] if len(row) > 2 else None
        total_amount = row[4] if len(row) > 4 else None

        if first == "INCOME":
            section = "income"
        elif first == "TOTAL INCOME":
            result["total_income"] = total_amount
            section = None
        elif first == "ADMINISTRATIVE EXPENSES":
            section = "expenses"
        elif first == "TOTAL ADMINISTRATIVE EXPENSES":
            result["total_expenses"] = total_amount
            section = None
        elif first == "OPERATING PROFIT":
            result["operating_profit"] = total_amount
        elif first == "FINANCE COSTS":
            section = "finance"
        elif first == "TOTAL FINANCE COSTS":
            result["total_finance_costs"] = total_amount
            section = None
        elif first == "PROFIT BEFORE TAX":
            result["profit_before_tax"] = total_amount
        elif first == "TAXATION":
            section = "tax"
        elif first == "TOTAL TAX":
            result["total_tax"] = total_amount
            section = None
        elif first.startswith("NET PROFIT"):
            result["net_profit"] = first
        elif first.startswith("NOTE:") or first.startswith("NOTE —"):
            result["notes"].append(first)
        elif first and section:
            item = {"label": first, "amount": line_amount}
            if section == "income":
                result["income_items"].append(item)
            elif section == "expenses":
                result["expense_items"].append(item)
            elif section == "finance":
                result["finance_cost_items"].append(item)
            elif section == "tax":
                result["tax_items"].append(item)

    return result


def _parse_fp(ws) -> dict:
    rows = _rows(ws)
    result: dict[str, Any] = {
        "company": rows[0][0],
        "title": rows[1][0],
        "as_at": rows[2][0],
        "non_current_asset_items": [],
        "total_non_current_assets": None,
        "current_asset_items": [],
        "total_current_assets": None,
        "total_assets": None,
        "equity_items": [],
        "total_equity": None,
        "non_current_liability_items": [],
        "total_liabilities": None,
        "total_equity_and_liabilities": None,
        "notes": [],
    }

    section: str | None = None
    for row in rows[4:]:
        first = str(row[0]).strip() if row[0] else ""
        line_amount = row[2] if len(row) > 2 else None
        total_amount = row[4] if len(row) > 4 else None

        if first == "NON-CURRENT ASSETS":
            section = "nca"
        elif first == "TOTAL NON-CURRENT ASSETS":
            result["total_non_current_assets"] = total_amount
            section = None
        elif first == "CURRENT ASSETS":
            section = "ca"
        elif first == "TOTAL CURRENT ASSETS":
            result["total_current_assets"] = total_amount
            section = None
        elif first.startswith("TOTAL ASSETS"):
            result["total_assets"] = first
        elif first == "EQUITY":
            section = "equity"
        elif first == "TOTAL EQUITY":
            result["total_equity"] = total_amount
            section = None
        elif first == "LIABILITIES":
            section = "liabilities"
        elif first == "NON-CURRENT LIABILITIES":
            pass
        elif first.startswith("TOTAL LIABILITIES"):
            result["total_liabilities"] = total_amount
            section = None
        elif first.startswith("TOTAL EQUITY AND LIABILITIES"):
            result["total_equity_and_liabilities"] = first
        elif first.startswith("NOTE"):
            result["notes"].append(first)
        elif first and section:
            item = {"label": first, "amount": line_amount}
            if section == "nca":
                result["non_current_asset_items"].append(item)
            elif section == "ca":
                result["current_asset_items"].append(item)
            elif section == "equity":
                result["equity_items"].append(item)
            elif section == "liabilities":
                result["non_current_liability_items"].append(item)

    return result


def _parse_notes(ws) -> list[dict]:
    rows = _rows(ws)
    notes: list[dict] = []
    current_note: dict | None = None

    for row in rows[4:]:
        first = str(row[0]).strip() if row[0] else ""
        col2 = row[1] if len(row) > 1 else None

        if first.startswith("Note ") and "—" in first:
            if current_note:
                notes.append(current_note)
            current_note = {"title": first, "lines": [], "table": []}
        elif current_note is not None:
            non_none = [(i, v) for i, v in enumerate(row) if v is not None]
            if len(non_none) >= 3:
                current_note["table"].append([v for v in row if v is not None])
            elif first:
                current_note["lines"].append({"label": first, "value": col2})

    if current_note:
        notes.append(current_note)
    return notes


def _parse_loan(ws) -> dict:
    rows = _rows(ws)
    result: dict[str, Any] = {
        "title": rows[0][0],
        "subtitle": rows[1][0],
        "period": rows[2][0],
        "columns": [],
        "rows": [],
        "summary": [],
    }

    header_found = False
    for row in rows[4:]:
        cells = [v for v in row if v is not None]
        if not cells:
            continue
        first = str(row[0]).strip() if row[0] else ""

        if not header_found and first == "Month":
            result["columns"] = [str(v) for v in row if v is not None]
            header_found = True
        elif header_found and first.startswith("Month"):
            result["rows"].append(list(row[:6]))
        elif header_found and first == "TOTAL":
            result["rows"].append(list(row[:6]))
        elif header_found and first:
            result["summary"].append({"label": first, "value": row[1] if len(row) > 1 else None})

    return result


@router.get("/")
def get_financial_statements():
    """Return parsed financial statements from the MLS Excel file."""
    if not _FS_FILE.exists():
        raise HTTPException(404, "Financial statements file not found on server")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(_FS_FILE), data_only=True)
        return {
            "profit_loss": _parse_pl(wb["Profit or Loss"]),
            "financial_position": _parse_fp(wb["Financial Position"]),
            "notes": _parse_notes(wb["Notes"]),
            "loan_schedule": _parse_loan(wb["Cooperative Loan Schedule"]),
        }
    except Exception as exc:
        raise HTTPException(500, f"Failed to parse financial statements: {exc}") from exc


@router.get("/download")
def download_financial_statements():
    """Download the original Excel financial statements file."""
    if not _FS_FILE.exists():
        raise HTTPException(404, "Financial statements file not found on server")
    return FileResponse(
        path=str(_FS_FILE),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="MLS_Financial_Statements.xlsx",
    )


# ── Category maps ─────────────────────────────────────────────────────────────
_INCOME_LABEL: dict[str, str] = {
    "School Fees": "Revenue — School Fees (Education Sector)",
    "Exam": "Revenue — Examination Fees",
    "Books": "Revenue — Books & Materials",
    "Stationery": "Revenue — Stationery Sales",
    "Uniform": "Revenue — Uniform Sales",
    "Freelance": "Revenue — Freelance / Professional Services",
    "Software": "Revenue — Software Services",
    "Business": "Revenue — Business Activities",
    "Investment": "Other Income — Investment / Interest",
    "Gift": "Other Income — Gifts Received",
    "Refund": "Other Income — Refunds",
}
_NON_OPERATING = {"Investment", "Gift", "Refund"}
_IGNORE_INCOME = {"Fund from Director", "Internal Transfer"}
_FINANCE_COST_CATS = {"Interest Expense", "Bank Charges"}
# Loan repayments reduce a balance sheet liability — excluded from P&L entirely (same as CIT treatment)
_IGNORE_EXPENSE = {"Internal Transfer", "Loans"}


def _run_computation(db: Session, start_date: str, end_date: str, company: str) -> dict:
    """Core computation shared by GET /computed and GET /computed/download."""
    sd = date.fromisoformat(start_date)
    ed = date.fromisoformat(end_date)

    # ── Lean transaction query — fetch only needed columns ────────────────────
    rows = db.execute(
        select(Transaction.type, Transaction.amount, Transaction.category)
        .where(Transaction.type.in_(["income", "expense"]))
        .where(Transaction.date >= sd)
        .where(Transaction.date <= ed)
    ).all()
    tx_count = len(rows)

    income_by_cat: dict[str, float] = {}
    expense_by_cat: dict[str, float] = {}
    for tx_type, tx_amount, tx_cat in rows:
        cat = tx_cat or "Other"
        if tx_type == "income":
            income_by_cat[cat] = income_by_cat.get(cat, 0) + tx_amount
        else:
            expense_by_cat[cat] = expense_by_cat.get(cat, 0) + tx_amount

    # ── Single-pass asset computation (avoids 5+ redundant loops) ────────────
    all_assets = db.query(Asset).all()
    asset_period_dep = 0.0
    asset_nbv_by_cat: dict[str, float] = {}
    active_asset_count = 0
    ppe_items: list[dict] = []

    for a in all_assets:
        dep_p = round(calc_period_dep(a, sd, ed), 2)
        asset_period_dep += dep_p

        if a.purchase_date > ed:
            continue  # not yet acquired at the reporting date

        active_asset_count += 1
        acc = round(calc_accumulated(a, ed), 2)
        nbv = round(a.purchase_cost - acc, 2)
        asset_nbv_by_cat[a.category] = round(asset_nbv_by_cat.get(a.category, 0) + nbv, 2)
        ppe_items.append({
            "name": a.name,
            "category": a.category,
            "purchase_date": str(a.purchase_date),
            "cost": a.purchase_cost,
            "useful_life_years": a.useful_life_years,
            "residual_value": a.residual_value,
            "depreciation_method": "SL" if a.depreciation_method == "straight_line" else "RB",
            "period_depreciation": dep_p,
            "accumulated_depreciation": acc,
            "nbv": nbv,
        })

    asset_period_dep = round(asset_period_dep, 2)
    if asset_period_dep > 0:
        expense_by_cat["Depreciation"] = round(
            expense_by_cat.get("Depreciation", 0) + asset_period_dep, 2
        )
    total_asset_nbv = round(sum(asset_nbv_by_cat.values()), 2)

    # ── Staff Loans Receivable ────────────────────────────────────────────────
    active_staff_loans = db.query(StaffLoan).filter(StaffLoan.is_active == True).all()
    staff_loan_items: list[dict] = []
    for sl in sorted(active_staff_loans, key=lambda x: x.employee_name):
        bal = _staff_loan_outstanding(sl)
        if bal > 0:
            staff_loan_items.append({"label": f"Staff Loan — {sl.employee_name}", "amount": bal})
    total_staff_loans = round(sum(i["amount"] for i in staff_loan_items), 2)

    # ── P&L — Income ─────────────────────────────────────────────────────────
    operating_income_items: list[dict] = []
    other_income_items: list[dict] = []
    for cat, amount in sorted(income_by_cat.items()):
        if cat in _IGNORE_INCOME:
            continue
        label = _INCOME_LABEL.get(cat, f"Revenue — {cat}")
        if cat in _NON_OPERATING:
            other_income_items.append({"label": label, "amount": round(amount, 2)})
        else:
            operating_income_items.append({"label": label, "amount": round(amount, 2)})

    total_operating = sum(i["amount"] for i in operating_income_items)
    total_other = sum(i["amount"] for i in other_income_items)
    total_income = round(total_operating + total_other, 2)

    # ── P&L — Expenses ───────────────────────────────────────────────────────
    admin_items: list[dict] = []
    finance_items: list[dict] = []
    for cat, amount in sorted(expense_by_cat.items()):
        if cat in _IGNORE_EXPENSE:
            continue
        if cat in _FINANCE_COST_CATS:
            finance_items.append({"label": cat, "amount": round(amount, 2)})
        else:
            admin_items.append({"label": cat, "amount": round(amount, 2)})

    total_admin = round(sum(i["amount"] for i in admin_items), 2)
    total_finance = round(sum(i["amount"] for i in finance_items), 2)
    operating_profit = round(total_operating - total_admin, 2)
    profit_before_tax = round(operating_profit + total_other - total_finance, 2)

    # ── CIT computation ───────────────────────────────────────────────────────
    cit_nil = total_operating <= 25_000_000
    cit_amount = 0.0 if cit_nil else round(
        max(profit_before_tax, 0) * (0.20 if total_operating <= 100_000_000 else 0.30), 2
    )
    edt_amount = round(max(profit_before_tax, 0) * 0.005, 2)
    total_tax = round(cit_amount + edt_amount, 2)
    net_profit = round(profit_before_tax - total_tax, 2)

    tax_items = [
        {
            "label": f"Companies Income Tax — CIT ({'NIL; Small Company < ₦25M' if cit_nil else f'{20 if total_operating <= 100_000_000 else 30}%'})",
            "amount": "NIL" if cit_nil else cit_amount,
        },
        {
            "label": f"Education Development Tax — EDT (0.5% × ₦{profit_before_tax:,.0f})",
            "amount": edt_amount,
        },
    ]

    pl_notes: list[str] = []
    if finance_items:
        pl_notes.append(
            f"NOTE — Bank Charges and Interest Expense (₦{total_finance:,.2f}) are classified "
            "under 'Finance Costs' per accounting standards, not in Administrative Expenses. "
            f"Total cash outflows = Admin Expenses ₦{total_admin:,.2f} + "
            f"Finance Costs ₦{total_finance:,.2f} = ₦{total_admin + total_finance:,.2f}. "
            "Loan repayments are balance sheet items (liability reduction) and are excluded from the income statement."
        )
    if asset_period_dep > 0:
        pl_notes.append(
            f"NOTE — Depreciation of ₦{asset_period_dep:,.2f} is computed from the asset register "
            "using the straight-line / reducing-balance method and included in Administrative Expenses."
        )

    period_label = f"{sd.strftime('%d %B %Y')} to {ed.strftime('%d %B %Y')}"

    profit_loss = {
        "company": company,
        "title": "Statement of Profit or Loss and Other Comprehensive Income",
        "period": f"For the Period: {period_label}",
        "income_items": operating_income_items + other_income_items,
        "total_income": total_income,
        "expense_items": admin_items,
        "total_expenses": total_admin,
        "operating_profit": operating_profit,
        "finance_cost_items": finance_items,
        "total_finance_costs": total_finance,
        "profit_before_tax": profit_before_tax,
        "tax_items": tax_items,
        "total_tax": total_tax,
        "net_profit": (
            f"NET PROFIT FOR THE PERIOD:  ₦{net_profit:,.0f}"
            if net_profit >= 0
            else f"NET LOSS FOR THE PERIOD:  ₦{abs(net_profit):,.0f}"
        ),
        "notes": pl_notes,
    }

    # ── Financial Position ────────────────────────────────────────────────────
    total_cash = round(
        sum(amt for t, amt, _ in rows if t == "income")
        - sum(amt for t, amt, _ in rows if t == "expense"),
        2,
    )
    loan_inflows = income_by_cat.get("Loans", 0)
    director_funds = income_by_cat.get("Fund from Director", 0)

    nca_items = [
        {"label": f"{cat} — Net Book Value", "amount": round(nbv, 2)}
        for cat, nbv in sorted(asset_nbv_by_cat.items())
        if nbv > 0
    ]
    total_current_assets = round(total_cash + total_staff_loans, 2)
    total_assets_value = round(total_asset_nbv + max(total_current_assets, 0), 2)

    current_asset_items: list[dict] = [
        {"label": "Net Cash Position (income – expenses for period)", "amount": total_cash},
    ]
    current_asset_items.extend(staff_loan_items)

    financial_position = {
        "company": company,
        "title": "Statement of Financial Position (Simplified)",
        "as_at": f"As at {ed.strftime('%d %B %Y')}",
        "non_current_asset_items": nca_items,
        "total_non_current_assets": total_asset_nbv,
        "current_asset_items": current_asset_items,
        "total_current_assets": total_current_assets,
        "total_assets": f"TOTAL ASSETS  ₦{total_assets_value:,.0f}",
        "equity_items": [
            {"label": "Retained Earnings — Net Profit for the Period", "amount": net_profit},
        ],
        "total_equity": net_profit,
        "non_current_liability_items": (
            [{"label": "Loans Received", "amount": round(loan_inflows, 2)}] if loan_inflows else []
        ) + (
            [{"label": "Director's Loan / Capital Injection", "amount": round(director_funds, 2)}] if director_funds else []
        ),
        "total_liabilities": round(loan_inflows + director_funds, 2),
        "total_equity_and_liabilities": "NOTE: Balance sheet is computed from transaction data for the selected period only.",
        "notes": [
            f"NOTE — Computed from {tx_count} transactions for {period_label}. "
            f"Non-current assets (₦{total_asset_nbv:,.2f} NBV) are from the asset register. "
            "Opening balances and accruals not captured as transactions are excluded."
        ],
    }

    # ── Statement of Cash Flows (Indirect Method) ─────────────────────────────
    period_asset_purchases = round(sum(
        a.purchase_cost for a in all_assets
        if sd <= a.purchase_date <= ed
    ), 2)

    op_cash_items = [{"label": "Net profit for the period", "amount": net_profit}]
    if asset_period_dep > 0:
        op_cash_items.append({"label": "Add: Depreciation (non-cash charge)", "amount": asset_period_dep})
    cash_from_ops = round(net_profit + asset_period_dep, 2)

    cf_investing: list[dict] = []
    if period_asset_purchases > 0:
        cf_investing.append({
            "label": "Purchase of property, plant and equipment",
            "amount": -period_asset_purchases,
        })
    total_investing = round(-period_asset_purchases, 2)

    cf_financing: list[dict] = []
    if loan_inflows > 0:
        cf_financing.append({"label": "Proceeds from borrowings (loans received)", "amount": round(loan_inflows, 2)})
    if director_funds > 0:
        cf_financing.append({"label": "Capital injection from director", "amount": round(director_funds, 2)})
    total_financing = round(loan_inflows + director_funds, 2)

    net_cash_movement = round(cash_from_ops + total_investing + total_financing, 2)

    cf_notes = [
        "Opening cash balance is not captured in this system and is stated as ₦0. For a complete reconciliation, add the opening bank balance from the prior period.",
    ]
    if period_asset_purchases > 0:
        cf_notes.append(
            f"Asset purchases of ₦{period_asset_purchases:,.2f} are sourced from the asset register and presented as investing outflows."
        )

    cash_flow_statement = {
        "company": company,
        "title": "Statement of Cash Flows",
        "period": f"For the Period: {period_label}",
        "method": "Indirect Method",
        "operating_items": op_cash_items,
        "total_operating": cash_from_ops,
        "investing_items": cf_investing,
        "total_investing": total_investing,
        "financing_items": cf_financing,
        "total_financing": total_financing,
        "net_movement": net_cash_movement,
        "opening_cash": 0.0,
        "closing_cash": net_cash_movement,
        "notes": cf_notes,
    }

    # ── Statement of Changes in Equity ────────────────────────────────────────
    changes_in_equity = {
        "company": company,
        "title": "Statement of Changes in Equity",
        "period": f"For the Period Ended {ed.strftime('%d %B %Y')}",
        "opening_retained_earnings": 0.0,
        "net_profit_for_period": net_profit,
        "dividends_drawings": 0.0,
        "closing_retained_earnings": net_profit,
        "note": (
            "Opening retained earnings represent accumulated profits from prior periods. "
            "These are not tracked in this system; the opening balance is stated as ₦0. "
            "For a complete statement, the prior period closing retained earnings should be used as the opening balance."
        ),
    }

    # ── Notes to Financial Statements ─────────────────────────────────────────
    cit_rate_desc = (
        "The company qualifies as a Small Company (annual turnover ≤ ₦25 million); CIT rate: 0% (NIL)."
        if cit_nil else
        f"Applicable CIT rate: {'20%' if total_operating <= 100_000_000 else '30%'} "
        f"({'Medium' if total_operating <= 100_000_000 else 'Large'} Company per CITA tiers)."
    )

    note1 = {
        "number": "1",
        "title": "Basis of Preparation and Significant Accounting Policies",
        "subsections": [
            {
                "subtitle": "1.1  Basis of Preparation",
                "text": (
                    f"These financial statements have been prepared on a going concern basis for {company} "
                    f"for the period {period_label}, expressed in Nigerian Naira (₦). "
                    "They are prepared in accordance with the International Financial Reporting Standard for "
                    "Small and Medium-sized Entities (IFRS for SMEs) as adopted by the Financial Reporting "
                    "Council of Nigeria (FRCN). These are unaudited management accounts compiled from transaction records and the asset register."
                ),
            },
            {
                "subtitle": "1.2  Revenue Recognition",
                "text": (
                    "Revenue is recognised when services are rendered or goods are delivered. "
                    "School fees and tuition income are recognised over the academic period to which they relate. "
                    "Examination fees, book and uniform sales are recognised at point of delivery."
                ),
            },
            {
                "subtitle": "1.3  Property, Plant and Equipment",
                "text": (
                    "Property, plant and equipment are measured at cost less accumulated depreciation and any impairment losses. "
                    "Depreciation is computed on the straight-line or reducing-balance method (as specified per asset) "
                    "over the estimated useful lives of the assets. "
                    "Assets are reviewed annually for impairment."
                ),
            },
            {
                "subtitle": "1.4  Income Tax",
                "text": (
                    "Income tax represents the estimated Companies Income Tax (CIT) computed under the "
                    "Companies Income Tax Act (CITA) Cap C21 LFN 2004 as amended by the Finance Acts. "
                    + cit_rate_desc +
                    " Education Development Tax (EDT) of 0.5% of assessable profit applies under the "
                    "Tertiary Education Trust Fund (Establishment) Act."
                ),
            },
            {
                "subtitle": "1.5  Cash and Cash Equivalents",
                "text": (
                    "Cash and cash equivalents comprise cash on hand and bank balances. "
                    "In these management accounts, the net cash position represents total income received "
                    "less total expenses paid during the period, derived from transaction records."
                ),
            },
        ],
    }

    note2 = {
        "number": "2",
        "title": "Revenue and Other Income",
        "text": "Revenue is disaggregated by category as follows:",
        "items": [{"label": i["label"], "amount": i["amount"]} for i in operating_income_items + other_income_items],
        "total": total_income,
        "subsections": [],
    }

    note3 = {
        "number": "3",
        "title": "Administrative Expenses",
        "text": "Administrative expenses are analysed as follows:",
        "items": [{"label": i["label"], "amount": i["amount"]} for i in admin_items],
        "total": total_admin,
        "subsections": [],
    }

    note4 = {
        "number": "4",
        "title": "Property, Plant and Equipment",
        "text": "A schedule of property, plant and equipment is set out below:",
        "items": ppe_items,
        "total_cost": round(sum(i["cost"] for i in ppe_items), 2),
        "total_period_dep": round(sum(i["period_depreciation"] for i in ppe_items), 2),
        "total_accumulated_dep": round(sum(i["accumulated_depreciation"] for i in ppe_items), 2),
        "total_nbv": round(sum(i["nbv"] for i in ppe_items), 2),
        "subsections": [],
    }

    fs_notes = [note1, note2, note3, note4]

    # ── Director's Declaration ─────────────────────────────────────────────────
    director_declaration = {
        "company": company,
        "period_ended": ed.strftime("%d %B %Y"),
        "clauses": [
            f"In the directors' opinion, there are reasonable grounds to believe that {company} will be able to pay its debts as and when they become due and payable.",
            "In the directors' opinion, the attached financial statements give a true and fair view of the financial position of the company as at the reporting date and of its performance for the period then ended, in accordance with IFRS for SMEs as adopted in Nigeria.",
            "The directors have considered the going concern status of the company and are satisfied that it will continue to operate as a going concern for the foreseeable future.",
        ],
        "signatories": [
            {"role": "Managing Director / CEO", "name": "", "date": ""},
            {"role": "Finance Director / CFO", "name": "", "date": ""},
        ],
        "authority": "This declaration is made in accordance with Sections 405–407 of the Companies and Allied Matters Act (CAMA) 2020.",
    }

    # ── Auditor's / Compilation Report ────────────────────────────────────────
    auditor_report = {
        "status": "unaudited",
        "type": "Compilation Report — Unaudited Management Accounts",
        "company": company,
        "period_ended": ed.strftime("%d %B %Y"),
        "paragraphs": [
            {
                "heading": "Status of These Accounts",
                "text": (
                    f"These financial statements of {company} for the period ended {ed.strftime('%d %B %Y')} "
                    "have been compiled by management from transaction records and the asset register. "
                    "They have NOT been subjected to an independent statutory audit."
                ),
                "list": [],
            },
            {
                "heading": "Requirements for Statutory Submission (CAC / FRCN / Regulatory Bodies)",
                "text": "For formal submission, these accounts must be:",
                "list": [
                    "Audited by an independent firm registered with the Institute of Chartered Accountants of Nigeria (ICAN) or recognised by the Association of Chartered Certified Accountants (ACCA).",
                    "Accompanied by an audit report (unqualified or qualified with explanation) in the prescribed format.",
                    "Signed by at least two directors and the company secretary.",
                    "Filed within the statutory period — within 6 months of financial year-end for FIRS; 42 days after AGM for CAC.",
                ],
            },
            {
                "heading": "FIRS CIT Return (Ready to File)",
                "text": (
                    "The Companies Income Tax (CIT) computation and pre-filled FIRS CIT Excel template "
                    "generated by this system are suitable for filing with the Federal Inland Revenue Service (FIRS). "
                    "The export includes revenue by sector, administrative expenses, non-allowable add-backs, "
                    "and the final CIT liability in the prescribed FIRS template format."
                ),
                "list": [],
            },
        ],
        "disclaimer": (
            "No independent auditor has expressed an opinion on these financial statements. "
            "These accounts are for internal management reporting and FIRS tax filing purposes only."
        ),
    }

    # ── Comparative figures (prior period, same duration shifted back 1 year) ──
    try:
        prior_sd = date(sd.year - 1, sd.month, sd.day)
        prior_ed = date(ed.year - 1, ed.month, ed.day)
    except ValueError:
        prior_sd = sd.replace(year=sd.year - 1, day=28)
        prior_ed = ed.replace(year=ed.year - 1, day=28)

    prior_rows = db.execute(
        select(Transaction.type, Transaction.amount, Transaction.category)
        .where(Transaction.type.in_(["income", "expense"]))
        .where(Transaction.date >= prior_sd)
        .where(Transaction.date <= prior_ed)
    ).all()
    p_inc: dict[str, float] = {}
    p_exp: dict[str, float] = {}
    for tx_type, tx_amount, tx_cat in prior_rows:
        cat = tx_cat or "Other"
        if tx_type == "income":
            p_inc[cat] = p_inc.get(cat, 0) + tx_amount
        else:
            p_exp[cat] = p_exp.get(cat, 0) + tx_amount

    # Reuse already-loaded assets (no second DB query for comparatives)
    p_dep = round(sum(calc_period_dep(a, prior_sd, prior_ed) for a in all_assets), 2)
    if p_dep > 0:
        p_exp["Depreciation"] = p_exp.get("Depreciation", 0) + p_dep

    p_total_inc = round(sum(v for k, v in p_inc.items() if k not in _IGNORE_INCOME), 2)
    p_admin = round(sum(v for k, v in p_exp.items() if k not in _FINANCE_COST_CATS and k not in _IGNORE_EXPENSE), 2)
    p_finance = round(sum(v for k, v in p_exp.items() if k in _FINANCE_COST_CATS), 2)
    p_op = round(p_total_inc - p_admin, 2)
    p_pbt = round(p_op - p_finance, 2)
    p_cit_nil = p_total_inc <= 25_000_000
    p_cit = 0.0 if p_cit_nil else round(max(p_pbt, 0) * (0.20 if p_total_inc <= 100_000_000 else 0.30), 2)
    p_edt = round(max(p_pbt, 0) * 0.005, 2)
    p_net = round(p_pbt - (p_cit + p_edt), 2)

    p_nbv_by_cat: dict[str, float] = {}
    for a in all_assets:
        if a.purchase_date > prior_ed:
            continue
        acc = calc_accumulated(a, prior_ed)
        p_nbv_by_cat[a.category] = round(p_nbv_by_cat.get(a.category, 0) + a.purchase_cost - acc, 2)
    p_total_nbv = round(sum(p_nbv_by_cat.values()), 2)

    comparatives = {
        "period": f"{prior_sd.strftime('%d %B %Y')} to {prior_ed.strftime('%d %B %Y')}",
        "total_income": p_total_inc,
        "total_admin_expenses": p_admin,
        "total_finance_costs": p_finance,
        "operating_profit": p_op,
        "profit_before_tax": p_pbt,
        "total_tax": round(p_cit + p_edt, 2),
        "net_profit": p_net,
        "total_asset_nbv": p_total_nbv,
        "transaction_count": len(prior_rows),
    }

    return {
        "mode": "computed",
        "period": period_label,
        "start_date": start_date,
        "end_date": end_date,
        "transaction_count": tx_count,
        "asset_count": active_asset_count,
        "total_asset_nbv": total_asset_nbv,
        "loan_repayments": round(expense_by_cat.get("Loans", 0), 2),
        "loan_drawdowns": round(loan_inflows, 2),
        "staff_loans_outstanding": total_staff_loans,
        "profit_loss": profit_loss,
        "financial_position": financial_position,
        "cash_flow_statement": cash_flow_statement,
        "changes_in_equity": changes_in_equity,
        "notes": fs_notes,
        "director_declaration": director_declaration,
        "auditor_report": auditor_report,
        "comparatives": comparatives,
    }


@router.get("/computed")
def get_computed_statements(
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    company: Optional[str] = Query("School Financial Statements", description="Entity name"),
    db: Session = Depends(get_db),
):
    """Compute P&L and balance sheet from transactions + asset register for a custom period."""
    co = company or "School Financial Statements"
    cache_key = f"{start_date}:{end_date}:{co}"
    if cached := _cache_get(cache_key):
        return JSONResponse(content=cached, headers={"Cache-Control": "max-age=10, stale-while-revalidate=20"})
    result = _run_computation(db, start_date, end_date, co)
    _cache_set(cache_key, result)
    return JSONResponse(content=result, headers={"Cache-Control": "max-age=10, stale-while-revalidate=20"})


@router.get("/computed/download")
def download_computed_statements(
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    company: Optional[str] = Query("School Financial Statements"),
    db: Session = Depends(get_db),
):
    """Download the computed financial statements as a formatted Excel workbook."""
    data = _run_computation(db, start_date, end_date, company or "School Financial Statements")
    buf = _generate_excel(data)
    filename = f"Financial_Statements_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Excel generation ──────────────────────────────────────────────────────────

def _generate_excel(data: dict) -> io.BytesIO:
    pl  = data["profit_loss"]
    fp  = data["financial_position"]
    cf  = data["cash_flow_statement"]
    ce  = data["changes_in_equity"]
    notes_data = data["notes"]
    decl = data["director_declaration"]
    audit = data["auditor_report"]
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    DARK_BLUE = "1F4E79"
    LIGHT_BLUE = "D6E4F0"
    LIGHT_GRAY = "F2F2F2"
    GREEN_BG = "E8F5E9"
    RED_BG = "FFEBEE"
    AMBER_BG = "FFF3CD"

    def _fill(hex_color: str) -> PatternFill:
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def _font(bold: bool = False, size: int = 10, color: str = "000000", italic: bool = False) -> Font:
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    def _num_fmt(ws, cell_ref: str, value: float | str | None, bold: bool = False, color: str = "000000", bg: str | None = None) -> None:
        cell = ws[cell_ref]
        if isinstance(value, (int, float)):
            cell.value = value
            cell.number_format = '#,##0.00'
        else:
            cell.value = str(value) if value is not None else "—"
        cell.font = _font(bold=bold, color=color)
        cell.alignment = Alignment(horizontal='right')
        if bg:
            cell.fill = _fill(bg)

    def _build_pl_sheet(ws) -> None:
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 22
        row = [1]  # mutable counter

        def r() -> int:
            return row[0]

        def next_row() -> None:
            row[0] += 1

        def merged_cell(text: str, font_kwargs: dict, align: str = 'center', bg: str | None = None) -> None:
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = text
            c.font = Font(name="Calibri", **font_kwargs)
            c.alignment = Alignment(horizontal=align, vertical='center')
            if bg:
                c.fill = _fill(bg)
                ws[f'B{r()}'].fill = _fill(bg)
            next_row()

        def section_header(label: str) -> None:
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = label
            c.font = _font(bold=True, size=10, color=DARK_BLUE)
            c.fill = _fill(LIGHT_BLUE)
            c.alignment = Alignment(horizontal='left', indent=1)
            ws[f'B{r()}'].fill = _fill(LIGHT_BLUE)
            next_row()

        def line_item(label: str, amount: float | str | None, indent: bool = True,
                      bold: bool = False, label_color: str = "000000",
                      amt_color: str = "000000", bg: str | None = None) -> None:
            c = ws[f'A{r()}']
            c.value = label
            c.font = _font(bold=bold, color=label_color)
            c.alignment = Alignment(horizontal='left', indent=(2 if indent else 1))
            if bg:
                c.fill = _fill(bg)
            _num_fmt(ws, f'B{r()}', amount, bold=bold, color=amt_color, bg=bg)
            next_row()

        # Company header
        merged_cell(pl['company'], dict(bold=True, size=14, color=DARK_BLUE))
        merged_cell(pl['title'], dict(bold=True, size=11, color=DARK_BLUE))
        merged_cell(pl['period'], dict(italic=True, size=10))
        merged_cell('All amounts in Nigerian Naira (₦)', dict(italic=True, size=9, color='808080'))
        next_row()  # blank

        # Income
        section_header("INCOME")
        for item in pl['income_items']:
            line_item(item['label'], item['amount'])
        line_item("TOTAL INCOME", pl['total_income'], indent=False, bold=True,
                  amt_color="2E7D32", bg=GREEN_BG)
        next_row()

        # Admin expenses
        section_header("ADMINISTRATIVE EXPENSES")
        for item in pl['expense_items']:
            line_item(item['label'], item['amount'])
        line_item("TOTAL ADMINISTRATIVE EXPENSES", pl['total_expenses'], indent=False, bold=True,
                  amt_color="C62828", bg=RED_BG)
        next_row()

        # Operating profit
        op = pl['operating_profit']
        op_color = "2E7D32" if op >= 0 else "C62828"
        op_bg = GREEN_BG if op >= 0 else RED_BG
        line_item("OPERATING PROFIT", op, indent=False, bold=True, amt_color=op_color, bg=op_bg)
        next_row()

        # Finance costs
        if pl.get('finance_cost_items'):
            section_header("FINANCE COSTS")
            for item in pl['finance_cost_items']:
                line_item(item['label'], item['amount'])
            line_item("TOTAL FINANCE COSTS", pl['total_finance_costs'], indent=False, bold=True,
                      amt_color="C62828", bg=RED_BG)
            next_row()

        # Profit before tax
        pbt = pl['profit_before_tax']
        pbt_color = "2E7D32" if pbt >= 0 else "C62828"
        pbt_bg = GREEN_BG if pbt >= 0 else RED_BG
        line_item("PROFIT BEFORE TAX", pbt, indent=False, bold=True, amt_color=pbt_color, bg=pbt_bg)
        next_row()

        # Taxation
        section_header("TAXATION")
        for item in pl['tax_items']:
            amt = item['amount']
            line_item(item['label'], amt if isinstance(amt, (int, float)) else str(amt))
        line_item("TOTAL TAX", pl['total_tax'], indent=False, bold=True, bg=LIGHT_GRAY)
        next_row()

        # Net profit banner
        merged_cell(pl['net_profit'], dict(bold=True, size=12, color="1B5E20"), bg=GREEN_BG)
        next_row()

        # Notes
        for note in pl.get('notes', []):
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = note
            c.font = _font(italic=True, size=9, color="7B5800")
            c.fill = _fill(AMBER_BG)
            ws[f'B{r()}'].fill = _fill(AMBER_BG)
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r()].height = 48
            next_row()

    def _build_fp_sheet(ws) -> None:
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 22
        row = [1]

        def r() -> int:
            return row[0]

        def next_row() -> None:
            row[0] += 1

        def merged_cell(text: str, font_kwargs: dict, align: str = 'center', bg: str | None = None) -> None:
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = text
            c.font = Font(name="Calibri", **font_kwargs)
            c.alignment = Alignment(horizontal=align, vertical='center')
            if bg:
                c.fill = _fill(bg)
                ws[f'B{r()}'].fill = _fill(bg)
            next_row()

        def section_header(label: str) -> None:
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = label
            c.font = _font(bold=True, size=10, color=DARK_BLUE)
            c.fill = _fill(LIGHT_BLUE)
            c.alignment = Alignment(horizontal='left', indent=1)
            ws[f'B{r()}'].fill = _fill(LIGHT_BLUE)
            next_row()

        def line_item(label: str, amount: float | str | None, indent: bool = True,
                      bold: bool = False, amt_color: str = "000000", bg: str | None = None) -> None:
            c = ws[f'A{r()}']
            c.value = label
            c.font = _font(bold=bold)
            c.alignment = Alignment(horizontal='left', indent=(2 if indent else 1))
            if bg:
                c.fill = _fill(bg)
            _num_fmt(ws, f'B{r()}', amount, bold=bold, color=amt_color, bg=bg)
            next_row()

        # Company header
        merged_cell(fp['company'], dict(bold=True, size=14, color=DARK_BLUE))
        merged_cell(fp['title'], dict(bold=True, size=11, color=DARK_BLUE))
        merged_cell(fp['as_at'], dict(italic=True, size=10))
        merged_cell('All amounts in Nigerian Naira (₦)', dict(italic=True, size=9, color='808080'))
        next_row()

        # Non-current assets
        if fp.get('non_current_asset_items'):
            section_header("NON-CURRENT ASSETS")
            for item in fp['non_current_asset_items']:
                line_item(item['label'], item['amount'])
            line_item("TOTAL NON-CURRENT ASSETS", fp['total_non_current_assets'],
                      indent=False, bold=True, bg=LIGHT_GRAY)
            next_row()

        # Current assets
        section_header("CURRENT ASSETS")
        for item in fp['current_asset_items']:
            line_item(item['label'], item['amount'])
        line_item("TOTAL CURRENT ASSETS", fp['total_current_assets'],
                  indent=False, bold=True, bg=LIGHT_GRAY)
        next_row()

        # Total assets
        merged_cell(fp['total_assets'], dict(bold=True, size=12, color="0D47A1"), bg="E3F2FD")
        next_row()

        # Equity
        section_header("EQUITY")
        for item in fp.get('equity_items', []):
            line_item(item['label'], item['amount'])
        line_item("TOTAL EQUITY", fp['total_equity'], indent=False, bold=True,
                  amt_color="0D47A1", bg=LIGHT_GRAY)
        next_row()

        # Liabilities
        if fp.get('non_current_liability_items'):
            section_header("LIABILITIES")
            for item in fp['non_current_liability_items']:
                line_item(item['label'], item['amount'])
            line_item("TOTAL LIABILITIES", fp['total_liabilities'], indent=False, bold=True,
                      amt_color="C62828", bg=RED_BG)
            next_row()

        # Notes
        for note in fp.get('notes', []):
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = note
            c.font = _font(italic=True, size=9, color="666666")
            c.fill = _fill(LIGHT_GRAY)
            ws[f'B{r()}'].fill = _fill(LIGHT_GRAY)
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r()].height = 48
            next_row()

    def _build_cf_sheet(ws) -> None:
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 22
        row = [1]

        def r() -> int: return row[0]
        def next_row() -> None: row[0] += 1

        def merged_header(text: str, font_kwargs: dict, bg: str | None = None) -> None:
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = text
            c.font = Font(name="Calibri", **font_kwargs)
            c.alignment = Alignment(horizontal='center', vertical='center')
            if bg:
                c.fill = _fill(bg); ws[f'B{r()}'].fill = _fill(bg)
            next_row()

        def section_hdr(label: str) -> None:
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = label
            c.font = _font(bold=True, size=10, color=DARK_BLUE)
            c.fill = _fill(LIGHT_BLUE); ws[f'B{r()}'].fill = _fill(LIGHT_BLUE)
            c.alignment = Alignment(horizontal='left', indent=1)
            next_row()

        def line(label: str, amount: float | None, indent: bool = True, bold: bool = False,
                 amt_color: str = "000000", bg: str | None = None) -> None:
            c = ws[f'A{r()}']
            c.value = label
            c.font = _font(bold=bold)
            c.alignment = Alignment(horizontal='left', indent=2 if indent else 1)
            if bg: c.fill = _fill(bg)
            if amount is not None:
                _num_fmt(ws, f'B{r()}', amount, bold=bold, color=amt_color, bg=bg)
            next_row()

        merged_header(cf['company'], dict(bold=True, size=14, color=DARK_BLUE))
        merged_header(cf['title'], dict(bold=True, size=11, color=DARK_BLUE))
        merged_header(cf['period'], dict(italic=True, size=10))
        merged_header(f"Method: {cf['method']}", dict(italic=True, size=9, color='808080'))
        merged_header('All amounts in Nigerian Naira (₦)', dict(italic=True, size=9, color='808080'))
        next_row()

        section_hdr("OPERATING ACTIVITIES")
        for item in cf['operating_items']:
            line(item['label'], item['amount'])
        line("Net cash generated from operating activities", cf['total_operating'],
             indent=False, bold=True,
             amt_color="2E7D32" if cf['total_operating'] >= 0 else "C62828",
             bg=GREEN_BG if cf['total_operating'] >= 0 else RED_BG)
        next_row()

        if cf['investing_items']:
            section_hdr("INVESTING ACTIVITIES")
            for item in cf['investing_items']:
                line(item['label'], item['amount'], amt_color="C62828")
            line("Net cash used in investing activities", cf['total_investing'],
                 indent=False, bold=True, amt_color="C62828", bg=RED_BG)
            next_row()

        if cf['financing_items']:
            section_hdr("FINANCING ACTIVITIES")
            for item in cf['financing_items']:
                line(item['label'], item['amount'], amt_color="2E7D32")
            line("Net cash from financing activities", cf['total_financing'],
                 indent=False, bold=True, amt_color="2E7D32", bg=GREEN_BG)
            next_row()

        mv = cf['net_movement']
        merged_header(
            f"NET {'INCREASE' if mv >= 0 else 'DECREASE'} IN CASH  ₦{mv:,.0f}",
            dict(bold=True, size=12, color="1B5E20" if mv >= 0 else "C62828"),
            bg=GREEN_BG if mv >= 0 else RED_BG,
        )
        line("Cash and cash equivalents at beginning of period", cf['opening_cash'], indent=False)
        line("Cash and cash equivalents at end of period", cf['closing_cash'],
             indent=False, bold=True, bg=LIGHT_GRAY)
        next_row()

        for note in cf.get('notes', []):
            ws.merge_cells(f'A{r()}:B{r()}')
            c = ws[f'A{r()}']
            c.value = note
            c.font = _font(italic=True, size=9, color="7B5800")
            c.fill = _fill(AMBER_BG); ws[f'B{r()}'].fill = _fill(AMBER_BG)
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r()].height = 36
            next_row()

    def _build_ce_sheet(ws) -> None:
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 24
        row = [1]

        def r() -> int: return row[0]
        def next_row() -> None: row[0] += 1

        def h(text: str, fk: dict, bg: str | None = None, cols: str = 'A:C') -> None:
            ws.merge_cells(f'{cols.split(":")[0]}{r()}:{cols.split(":")[1]}{r()}')
            c = ws[f'A{r()}']
            c.value = text
            c.font = Font(name="Calibri", **fk)
            c.alignment = Alignment(horizontal='center')
            if bg:
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{r()}'].fill = _fill(bg)
            next_row()

        h(ce['company'], dict(bold=True, size=14, color=DARK_BLUE))
        h(ce['title'], dict(bold=True, size=11, color=DARK_BLUE))
        h(ce['period'], dict(italic=True, size=10))
        h('All amounts in Nigerian Naira (₦)', dict(italic=True, size=9, color='808080'))
        next_row()

        # Column headers
        for col, label in [('A', 'Description'), ('B', 'Retained Earnings (₦)'), ('C', 'Total (₦)')]:
            c = ws[f'{col}{r()}']
            c.value = label
            c.font = _font(bold=True, color="FFFFFF")
            c.fill = _fill(DARK_BLUE)
            c.alignment = Alignment(horizontal='center' if col != 'A' else 'left', indent=1)
        next_row()

        rows_ce = [
            ("Opening balance — 1 January", ce['opening_retained_earnings'], ce['opening_retained_earnings']),
            ("Net profit for the period", ce['net_profit_for_period'], ce['net_profit_for_period']),
            ("Dividends / drawings", ce['dividends_drawings'], ce['dividends_drawings']),
            ("Closing balance", ce['closing_retained_earnings'], ce['closing_retained_earnings']),
        ]
        for i, (lbl, re_val, tot) in enumerate(rows_ce):
            is_close = i == len(rows_ce) - 1
            ws[f'A{r()}'].value = lbl
            ws[f'A{r()}'].font = _font(bold=is_close)
            ws[f'A{r()}'].alignment = Alignment(horizontal='left', indent=1)
            for col, val in [('B', re_val), ('C', tot)]:
                cell = ws[f'{col}{r()}']
                cell.value = val
                cell.number_format = '#,##0.00'
                cell.font = _font(bold=is_close, color="2E7D32" if val >= 0 else "C62828")
                cell.alignment = Alignment(horizontal='right')
                if is_close:
                    cell.fill = _fill(LIGHT_GRAY)
            next_row()

        next_row()
        ws.merge_cells(f'A{r()}:C{r()}')
        c = ws[f'A{r()}']
        c.value = ce['note']
        c.font = _font(italic=True, size=9, color="7B5800")
        c.fill = _fill(AMBER_BG)
        ws[f'B{r()}'].fill = _fill(AMBER_BG); ws[f'C{r()}'].fill = _fill(AMBER_BG)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r()].height = 48

    def _build_notes_sheet(ws) -> None:
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        row = [1]

        def r() -> int: return row[0]
        def next_row() -> None: row[0] += 1

        def section(text: str, bold: bool = False, indent: int = 0, bg: str | None = None, wrap: bool = False) -> None:
            ws.merge_cells(f'A{r()}:E{r()}')
            c = ws[f'A{r()}']
            c.value = text
            c.font = _font(bold=bold, color=DARK_BLUE if bold and indent == 0 else "000000")
            c.alignment = Alignment(horizontal='left', indent=indent, wrap_text=wrap)
            if bg:
                for col in list('ABCDE'):
                    ws[f'{col}{r()}'].fill = _fill(bg)
            if wrap:
                ws.row_dimensions[r()].height = max(30, len(text) // 4)
            next_row()

        for note in notes_data:
            section(f"Note {note['number']} — {note['title']}", bold=True, bg=LIGHT_BLUE)

            if 'subsections' in note and note['subsections']:
                for sub in note['subsections']:
                    section(sub['subtitle'], bold=True, indent=1)
                    section(sub['text'], indent=2, wrap=True)
                    next_row()
            elif 'items' in note and note['items']:
                if note.get('text'):
                    section(note['text'], indent=1, wrap=True)
                next_row()

                if note['number'] == '4' and note['items']:
                    # PPE table
                    hdrs = ['Asset Name', 'Cost (₦)', 'Period Dep. (₦)', 'Accum. Dep. (₦)', 'NBV (₦)']
                    for ci, hdr in enumerate(hdrs):
                        col = chr(65 + ci)
                        c = ws[f'{col}{r()}']
                        c.value = hdr
                        c.font = _font(bold=True, color="FFFFFF", size=9)
                        c.fill = _fill(DARK_BLUE)
                        c.alignment = Alignment(horizontal='center' if ci > 0 else 'left', indent=1)
                    next_row()
                    for item in note['items']:
                        for ci, val in enumerate([
                            f"{item['name']} ({item['category']}, {item['depreciation_method']})",
                            item['cost'], item['period_depreciation'],
                            item['accumulated_depreciation'], item['nbv'],
                        ]):
                            col = chr(65 + ci)
                            c = ws[f'{col}{r()}']
                            if isinstance(val, float):
                                c.value = val; c.number_format = '#,##0.00'
                                c.alignment = Alignment(horizontal='right')
                            else:
                                c.value = val; c.alignment = Alignment(horizontal='left', indent=1)
                            c.font = _font(size=9)
                        next_row()
                    # Totals
                    for ci, (lbl, key) in enumerate([
                        ('TOTAL', None), ('total_cost', 'total_cost'),
                        ('total_period_dep', 'total_period_dep'),
                        ('total_accumulated_dep', 'total_accumulated_dep'),
                        ('total_nbv', 'total_nbv'),
                    ]):
                        col = chr(65 + ci)
                        c = ws[f'{col}{r()}']
                        if ci == 0:
                            c.value = 'TOTAL'
                        else:
                            c.value = note.get(key, 0)
                            c.number_format = '#,##0.00'
                            c.alignment = Alignment(horizontal='right')
                        c.font = _font(bold=True)
                        c.fill = _fill(LIGHT_GRAY)
                    next_row()
                else:
                    # Simple two-column table
                    for item in note['items']:
                        ws[f'A{r()}'].value = item['label']
                        ws[f'A{r()}'].font = _font(size=9)
                        ws[f'A{r()}'].alignment = Alignment(horizontal='left', indent=2)
                        _num_fmt(ws, f'B{r()}', item['amount'], bold=False)
                        next_row()
                    ws[f'A{r()}'].value = 'TOTAL'
                    ws[f'A{r()}'].font = _font(bold=True)
                    _num_fmt(ws, f'B{r()}', note.get('total', 0), bold=True, bg=LIGHT_GRAY)
                    ws[f'B{r()}'].fill = _fill(LIGHT_GRAY)
                    next_row()
            next_row()

    def _build_declaration_sheet(ws) -> None:
        ws.column_dimensions['A'].width = 90
        row = [1]

        def r() -> int: return row[0]
        def next_row() -> None: row[0] += 1

        def blk(text: str, bold: bool = False, size: int = 10, color: str = "000000",
                bg: str | None = None, wrap: bool = True, height: int = 24) -> None:
            ws.merge_cells(f'A{r()}:A{r()}')
            c = ws[f'A{r()}']
            c.value = text
            c.font = Font(name="Calibri", bold=bold, size=size, color=color)
            c.alignment = Alignment(wrap_text=wrap, indent=1, vertical='top')
            if bg: c.fill = _fill(bg)
            ws.row_dimensions[r()].height = height
            next_row()

        # ── Auditor / Compilation Report ────
        blk(audit['type'], bold=True, size=13, color=DARK_BLUE, height=20)
        blk(f"Company: {audit['company']}  |  Period ended: {audit['period_ended']}", size=10, color="555555", height=16)
        next_row()

        for para in audit['paragraphs']:
            blk(para['heading'], bold=True, size=11, color=DARK_BLUE, height=18)
            if para.get('text'):
                blk(para['text'], wrap=True, height=40)
            for li in para.get('list', []):
                blk(f"    • {li}", wrap=True, height=30)
            next_row()

        blk(audit['disclaimer'], bold=False, size=9, color="555555", bg=AMBER_BG, wrap=True, height=36)
        next_row(); next_row()

        # ── Director's Declaration ────
        blk("DIRECTORS' DECLARATION", bold=True, size=13, color=DARK_BLUE, height=20)
        blk(f"Company: {decl['company']}  |  Period ended: {decl['period_ended']}", size=10, color="555555", height=16)
        next_row()
        blk("The directors of the above-named company hereby declare that:", height=16)
        for i, clause in enumerate(decl['clauses'], 1):
            blk(f"({chr(96+i)})  {clause}", wrap=True, height=36)
        next_row()

        blk("Signed in accordance with a resolution of the Board of Directors:", height=16)
        next_row()
        for sig in decl['signatories']:
            blk(f"{sig['role']}:  ___________________________________    Name: ________________    Date: ____________", height=20)
        next_row()

        blk(decl['authority'], size=9, color="555555", bg=LIGHT_GRAY, wrap=True, height=30)

    ws_pl = wb.active
    ws_pl.title = "Profit or Loss"
    _build_pl_sheet(ws_pl)

    ws_fp = wb.create_sheet("Financial Position")
    _build_fp_sheet(ws_fp)

    ws_cf = wb.create_sheet("Cash Flows")
    _build_cf_sheet(ws_cf)

    ws_ce = wb.create_sheet("Changes in Equity")
    _build_ce_sheet(ws_ce)

    ws_notes = wb.create_sheet("Notes to FS")
    _build_notes_sheet(ws_notes)

    ws_decl = wb.create_sheet("Declaration & Audit")
    _build_declaration_sheet(ws_decl)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
