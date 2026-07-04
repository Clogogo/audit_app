"""
Nigeria CIT (Company Income Tax) — FIRS CITA computation and template export.

Aggregates transactions by category for a given year, computes the CIT
liability using Nigerian CITA tiers, and generates a pre-filled copy of
the official FIRS CIT Excel return template.
"""
import io
from datetime import date
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, Query
from utils.auth import require_permission
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract

from database import get_db
from models import Transaction, Asset
from routers.assets import calc_period_dep

router = APIRouter(prefix="/tax", tags=["tax"], dependencies=[Depends(require_permission("tax"))])


# ── Template path ──────────────────────────────────────────────────────────────
_TEMPLATE = Path(__file__).parent.parent.parent.parent / "2025_2522801422271_CIT.xlsx"

# ── Income category → CIT revenue sector ──────────────────────────────────────
# Sector key must match INCOME_STATEMENT col-C labels (case-insensitive match).
_INCOME_TO_SECTOR: dict[str, str] = {
    "School Fees":   "EDUCATION",
    "Exam":          "EDUCATION",
    "Books":         "EDUCATION",
    "Stationery":    "EDUCATION",
    "Uniform":       "EDUCATION",
    "Freelance":     "PROFESSIONAL, SCIENTIFIC AND TECHNICAL ACTIVITIES",
    "Software":      "PROFESSIONAL, SCIENTIFIC AND TECHNICAL ACTIVITIES",
    "Business":      "WHOLESALE AND RETAIL TRADE",
    # These flow to Other Income, not primary revenue
    "Investment":    None,
    "Gift":          None,
    "Refund":        None,
    # Non-revenue — director capital injection
    "Fund from Director": None,
    "Internal Transfer":  None,
    "Other":         "Others",
}

# Income category → OTHER_INCOME sheet label (col C)
_INCOME_TO_OTHER: dict[str, str] = {
    "Investment": "Interest Income",
    "Gift":       "Others",
    "Refund":     "Others",
}

# ── Expense category → (admin expenses template label, is_non_allowable) ──────
# Template labels are col-B values in ADMINISTRATIVE_EXPENSES sheet.
_EXPENSE_MAP: dict[str, tuple[str, bool]] = {
    "Salary and Wages":                           ("Salary and Wages", False),
    "IOU (Advance Salary)":                        ("Salary and Wages", False),
    "Employee Benefit Expenses":                   ("Employee Benefit Expenses", False),
    "Directors Emoluments":                        ("Directors Emoluments", False),
    "Directors Expenses":                          ("Directors expenses", False),
    "Staff Allowances":                            ("Staff Allowances", False),
    "Staff Loan Written Off":                      ("Staff Loan written off", False),
    "Staff Gratuity":                              ("Staff Gratuity", False),
    "Pension and Gratuity":                        ("Pension and Gratuity ", False),
    "Other Staff & Related Cost":                  ("Other staff & related Cost", False),
    "Depreciation":                                ("depreciation", True),
    "Amortisation of Intangible Assets":           ("Amortisation of intangible assetsX", False),
    "Government and Regulatory Costs":             ("Government and regulatory costs", False),
    "Advertisement & Promotion":                   ("Advertisement & Promotion", False),
    "Bad Debt Written Off":                        ("Bad Debt written off", False),
    "Bank Charges":                                ("Bank charges", False),
    "Interest Expense":                            ("Interest Expense", False),
    "Rooms/Accommodation Cost":                    ("Rooms/ Accomodation cost", False),
    "Audit Fees":                                  ("Audit fees ", False),
    "Donations":                                   ("Donations", True),
    "Entertainment Expenses":                      ("Entertainment expenses", True),
    "Postages":                                    ("Postages", False),
    "Fuel Expenses":                               ("Fuel expenses", False),
    "Generator Expenses":                          ("Generator expenses", False),
    "Head Office Cost":                            ("Head office cost", False),
    "Internet Cost":                               ("Internet Cost", False),
    "Loss on Sale of Property, Plant and Equipment": ("Loss on sale of property, plants and equipments", True),
    "Medical Expenses":                            ("Medical Expenses", False),
    "Newspaper & Periodicals":                     ("Newspaper & Periodicals", False),
    "Outsourcing Services Expenses":               ("Outsourcing Services expenses", False),
    "Rates":                                       ("Rates", False),
    "Rent & Rates":                                ("Rent & Rates", False),
    "Repairs and Maintenance":                     ("Repairs and Maintenance", False),
    "Research Cost":                               ("Research cost", False),
    "Security Expenses":                           ("Security expenses", False),
    "Selling & Distribution Expenses":             ("Selling & Distribution Expenses", False),
    "Service Charge":                              ("Service Charge", False),
    "Public Relation":                             ("Public Relation", False),
    "Publication Cost":                            ("Publication Cost", False),
    "Permits":                                     ("Permits", False),
    "Telephone":                                   ("Telephone", False),
    "Trade Fairs & Exhibitions":                   ("Trade fairs & exhibitions", False),
    "Training and Development":                    ("Training and Development", False),
    "Transport of Supplies":                       ("Transport of supplies", False),
    "Utilities":                                   ("Utilities", False),
    "Website Design & Maintenance":                ("Website Design & Maintenance", False),
    "Software Maintenance":                        ("Software maintenance", False),
    "Software":                                    ("Software maintenance", False),  # virtual account / POS subscription fees
    "Exam":                                        ("Stationeries expenses", False),  # exam registration & printing — educational materials
    "Books":                                       ("Stationeries expenses", False),  # educational books & materials
    "Cleaning & Janitorial Services":              ("Cleaning & Janitory Services", False),
    "Corporate Promotion":                         ("Corporate promotion ", False),
    "Freight & Transport Expenses":                ("Freight & Transport Expenses", False),
    "Subscription & Dues":                         ("Subscription & Dues", False),
    "Sponsorship":                                 ("Sponsorship ", False),
    "Stationeries Expenses":                       ("Stationeries expenses", False),
    "Other Professional Fees":                     ("other professional fees", False),
    "Fines":                                       ("Fines", True),
    "Penalties":                                   ("penalties", True),
    "AMCON Charges":                               ("AMCON charges", True),
    "Realized Foreign Exchange Loss":              ("Realized foreign exchange loss", False),
    "Corporate Social Responsibilities":           ("Corporate Social Responsibilities", True),
    "Royalties":                                   ("Royalties", False),
    "Electricity & Power":                         ("Electricity & Power", False),
    "PMS":                                         ("PMS", False),
    "Health, Safety and Environmental Expenses":   ("Health, safety and environmental expenses", False),
    "Licences & Leases Charges":                   ("Licences & leases charges", False),
    "Other Statutory Charges & Levies":            ("Other statutory charges & levies", False),
    "AGO-Diesel":                                  ("AGO-Diesel", False),
    "Travel Expenses":                             ("Travel Expenses", False),
    "Other":                                       ("Others", False),
    "Internal Transfer":                           None,  # type: ignore[assignment]
}

# ── Expense categories that should never appear — flag for user review ────────
# These are excluded from CIT calculations but surfaced as warnings in the summary.
_FLAGGED_EXPENSE_CATEGORIES: dict[str, str] = {
    "School Fees": (
        "School Fees recorded as an expense — a school should not pay school fees. "
        "Please review these transactions and re-categorise them correctly."
    ),
}

# ── PROFIT_ADJUSTMENT add-back rows (col D input) ─────────────────────────────
_ADDBACK_ROW: dict[str, int] = {
    "Donations":                                              15,
    "Entertainment Expenses":                                 18,
    "Loss on Sale of Property, Plant and Equipment":          17,
    "Fines":                                                  29,
    "Penalties":                                              29,  # merged with Fines in row 29
}

# ── PROFIT_ADJUSTMENT deduction rows (col D input) ────────────────────────────
_DEDUCTION_ROW: dict[str, int] = {
    "Investment": 38,  # Frank Investment Income (dividends/interest)
}


def _cit_rate(total_revenue: float) -> tuple[str, float]:
    """Return (company_size_label, cit_rate) per Nigerian CITA tiers."""
    if total_revenue <= 25_000_000:
        return "Small Company (≤ ₦25M)", 0.0
    if total_revenue <= 100_000_000:
        return "Medium Company (₦25M – ₦100M)", 0.20
    return "Large Company (> ₦100M)", 0.30


def _build_summary(
    db: Session,
    *,
    year: int | None = None,
    start_date: "date | None" = None,
    end_date: "date | None" = None,
) -> dict:
    q = db.query(Transaction).filter(Transaction.type.in_(["income", "expense"]))
    if start_date and end_date:
        q = q.filter(Transaction.date >= start_date, Transaction.date <= end_date)
    elif year:
        q = q.filter(extract("year", Transaction.date) == year)
    transactions = q.all()

    income_by_cat: dict[str, float] = {}
    expense_by_cat: dict[str, float] = {}
    for tx in transactions:
        cat = tx.category or "Other"
        if tx.type == "income":
            income_by_cat[cat] = income_by_cat.get(cat, 0) + tx.amount
        else:
            expense_by_cat[cat] = expense_by_cat.get(cat, 0) + tx.amount

    # Revenue by sector
    revenue_by_sector: dict[str, float] = {}
    other_income: dict[str, float] = {}
    for cat, amount in income_by_cat.items():
        sector = _INCOME_TO_SECTOR.get(cat)
        if sector:
            revenue_by_sector[sector] = revenue_by_sector.get(sector, 0) + amount
        elif cat in _INCOME_TO_OTHER:
            label = _INCOME_TO_OTHER[cat]
            other_income[label] = other_income.get(label, 0) + amount
        # Fund from Director / Internal Transfer → ignored (not taxable revenue)

    # ── Merge asset depreciation into expense_by_cat ─────────────────────────
    all_assets = db.query(Asset).all()
    if start_date and end_date:
        asset_dep = sum(calc_period_dep(a, start_date, end_date) for a in all_assets)
    elif year:
        _ys = date(year, 1, 1)
        _ye = date(year, 12, 31)
        asset_dep = sum(calc_period_dep(a, _ys, _ye) for a in all_assets)
    else:
        asset_dep = 0.0
    if asset_dep > 0:
        expense_by_cat["Depreciation"] = expense_by_cat.get("Depreciation", 0) + round(asset_dep, 2)

    # Flagged categories — excluded from CIT but surfaced as warnings
    flagged_categories: list[dict] = []
    for cat, amount in expense_by_cat.items():
        if cat in _FLAGGED_EXPENSE_CATEGORIES:
            flagged_categories.append({
                "category": cat,
                "amount": round(amount, 2),
                "reason": _FLAGGED_EXPENSE_CATEGORIES[cat],
            })

    # Admin expenses and add-backs
    admin_expenses: dict[str, float] = {}
    add_backs: dict[str, float] = {}
    for cat, amount in expense_by_cat.items():
        if cat in _FLAGGED_EXPENSE_CATEGORIES:
            continue  # excluded — user must re-categorise
        mapping = _EXPENSE_MAP.get(cat)
        if mapping is None:
            continue
        label, is_non_allowable = mapping
        admin_expenses[label] = admin_expenses.get(label, 0) + amount
        if is_non_allowable:
            add_backs[cat] = add_backs.get(cat, 0) + amount

    total_revenue = sum(revenue_by_sector.values())
    total_other_income = sum(other_income.values())
    total_admin_expenses = sum(admin_expenses.values())
    gross_profit = total_revenue - total_admin_expenses
    profit_before_tax = gross_profit + total_other_income
    total_add_backs = sum(add_backs.values())
    adjusted_profit = profit_before_tax + total_add_backs

    company_size, tax_rate = _cit_rate(total_revenue)
    tax_payable = max(0.0, round(adjusted_profit * tax_rate, 2)) if adjusted_profit > 0 else 0.0
    minimum_tax = round(total_revenue * 0.005, 2) if total_revenue > 0 else 0.0

    minimum_tax_applies = tax_rate > 0 and minimum_tax > tax_payable
    effective_tax = minimum_tax if minimum_tax_applies else tax_payable

    period_label = (
        f"{start_date} to {end_date}"
        if start_date and end_date
        else f"{year} (Jan – Dec)"
    )

    return {
        "year": year,
        "start_date": str(start_date) if start_date else (f"{year}-01-01" if year else None),
        "end_date": str(end_date) if end_date else (f"{year}-12-31" if year else None),
        "period_label": period_label,
        "revenue_by_sector": revenue_by_sector,
        "other_income": other_income,
        "admin_expenses": admin_expenses,
        "total_revenue": total_revenue,
        "total_other_income": total_other_income,
        "total_admin_expenses": total_admin_expenses,
        "gross_profit": gross_profit,
        "profit_before_tax": profit_before_tax,
        "add_backs": add_backs,
        "total_add_backs": total_add_backs,
        "adjusted_profit": adjusted_profit,
        "company_size": company_size,
        "tax_rate": tax_rate,
        "tax_payable": tax_payable,
        "minimum_tax": minimum_tax,
        "minimum_tax_applies": minimum_tax_applies,
        "effective_tax_payable": effective_tax,
        "transaction_count": len(transactions),
        "flagged_categories": flagged_categories,
    }


@router.get("/summary")
def get_tax_summary(
    year: Optional[int] = Query(None, description="Assessment year, e.g. 2025"),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    """Return CIT computation figures for a year or custom date range."""
    sd = date.fromisoformat(start_date) if start_date else None
    ed = date.fromisoformat(end_date) if end_date else None
    if not sd and not ed and not year:
        year = date.today().year - 1
    return _build_summary(db, year=year, start_date=sd, end_date=ed)


@router.get("/years")
def get_available_years(db: Session = Depends(get_db)):
    """Return list of years that have transactions, for the year picker."""
    from sqlalchemy import func as sqlfunc
    # extract() is portable across SQLite and Postgres; strftime() is SQLite-only.
    year_col = sqlfunc.extract("year", Transaction.date)
    rows = (
        db.query(year_col.label("yr"))
        .filter(Transaction.type.in_(["income", "expense"]))
        .distinct()
        .order_by(year_col.desc())
        .all()
    )
    return [int(r.yr) for r in rows if r.yr]


@router.get("/export")
def export_cit_excel(
    year: Optional[int] = Query(None, description="Assessment year, e.g. 2025"),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    share_capital: float = Query(100_000.0, description="Issued share capital (NGN)"),
    retained_earnings_bf: float = Query(0.0, description="Retained earnings brought forward from prior year"),
    short_term_debt: float = Query(0.0, description="Outstanding short-term loan balance not tracked as transactions"),
    other_reserves: float = Query(0.0, description="Other equity reserves (e.g. capital contribution above share capital)"),
    db: Session = Depends(get_db),
):
    """Generate and return a pre-filled FIRS CIT Excel return template."""
    from fastapi import HTTPException
    from openpyxl import load_workbook

    if not _TEMPLATE.exists():
        raise HTTPException(404, "CIT template file not found on server")

    sd = date.fromisoformat(start_date) if start_date else None
    ed = date.fromisoformat(end_date) if end_date else None
    if not sd and not ed and not year:
        year = date.today().year - 1

    summary = _build_summary(db, year=year, start_date=sd, end_date=ed)
    year = year or (sd.year if sd else date.today().year - 1)
    ye = date(year, 12, 31)

    wb = load_workbook(str(_TEMPLATE), keep_vba=False)

    # Ensure Excel recalculates all formulas when the file is opened
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # ── USER_INPUT: assessment year ────────────────────────────────────────────
    ws = wb["USER_INPUT"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "ACCOUNTING YEAR":
                ws.cell(row=cell.row, column=cell.column + 1).value = year

    # ── INCOME_STATEMENT: clear then write revenue by sector ──────────────────
    ws_inc = wb["INCOME_STATEMENT"]
    for r in range(7, 28):
        ws_inc.cell(row=r, column=4).value = None
    sector_row: dict[str, int] = {}
    for row in ws_inc.iter_rows(min_row=7, max_row=27):
        for cell in row:
            if cell.column == 3 and cell.value:
                sector_row[str(cell.value).strip().upper()] = cell.row

    for sector, amount in summary["revenue_by_sector"].items():
        row_num = sector_row.get(sector.upper())
        if row_num:
            ws_inc.cell(row=row_num, column=4).value = round(amount, 2)

    # P&L inputs that feed the already-existing formulas:
    #   E116 = D116 (Taxation)   E127 = D127 (Retained Earnings B/F)
    ws_inc["D116"] = 0                              # small company → 0% CIT
    ws_inc["D127"] = round(retained_earnings_bf, 2) # prior-year retained earnings

    # ── ADMINISTRATIVE_EXPENSES: clear then write expense line items ───────────
    ws_adm = wb["ADMINISTRATIVE_EXPENSES"]
    for r in range(4, 71):
        ws_adm.cell(row=r, column=3).value = None
    admin_row: dict[str, int] = {}
    for row in ws_adm.iter_rows(min_row=4, max_row=70):
        for cell in row:
            if cell.column == 2 and cell.value:
                admin_row[str(cell.value).strip().lower()] = cell.row

    for label, amount in summary["admin_expenses"].items():
        row_num = admin_row.get(label.strip().lower())
        if row_num:
            ws_adm.cell(row=row_num, column=3).value = round(amount, 2)

    # ── OTHER_INCOME: clear then write other income items ─────────────────────
    ws_oi = wb["OTHER_INCOME"]
    for r in range(4, 18):
        ws_oi.cell(row=r, column=4).value = None
    other_row: dict[str, int] = {}
    for row in ws_oi.iter_rows(min_row=4, max_row=17):
        for cell in row:
            if cell.column == 3 and cell.value:
                other_row[str(cell.value).strip().lower()] = cell.row

    for label, amount in summary["other_income"].items():
        row_num = other_row.get(label.strip().lower())
        if row_num:
            ws_oi.cell(row=row_num, column=4).value = round(amount, 2)

    # ── PROFIT_ADJUSTMENT: add-backs ──────────────────────────────────────────
    ws_pa = wb["PROFIT_ADJUSTMENT"]

    # E11 = =INCOME_STATEMENT!E115 (already a formula in template — no action needed)

    # E13 (depreciation): the template formula uses a cross-sheet structured
    # table reference that Excel may not resolve correctly, so write the
    # computed value directly to override it.
    dep_amount = summary["admin_expenses"].get("depreciation", 0)
    ws_pa["E13"] = round(dep_amount, 2)

    # Other non-allowable add-backs → col D (E33 = SUM(D14:D32)+E13)
    addback_values: dict[int, float] = {}
    for cat, amount in summary["add_backs"].items():
        if cat == "Depreciation":
            continue  # handled via E13 above
        if cat in ("Fines", "Penalties"):
            addback_values[29] = addback_values.get(29, 0) + amount
        else:
            row_num = _ADDBACK_ROW.get(cat)
            if row_num:
                addback_values[row_num] = addback_values.get(row_num, 0) + amount
    for row_num, amount in addback_values.items():
        ws_pa.cell(row=row_num, column=4).value = round(amount, 2)

    # Deductions (Frank Investment Income, etc.) → col D (E41 = SUM(D36:D40))
    for cat, row_num in _DEDUCTION_ROW.items():
        amt = summary["other_income"].get(
            _INCOME_TO_OTHER.get(cat, ""), 0
        )
        if amt:
            ws_pa.cell(row=row_num, column=4).value = round(amt, 2)

    # ── STATEMENT_OF_FINANCIAL_POSITION ───────────────────────────────────────
    ws_sfp = wb["STATEMENT_OF_FINANCIAL_POSITION"]

    # 1. PPE net book value — accumulated depreciation from purchase date to year end
    all_assets = db.query(Asset).all()
    ppe_nbv = sum(
        (a.purchase_cost or 0) - calc_period_dep(a, a.purchase_date, ye)
        for a in all_assets
    )
    ws_sfp["D8"] = round(max(0.0, ppe_nbv), 2)

    # 2. Cash at bank — net of all income and expense transactions for the period
    _start = sd or date(year, 1, 1)
    _end   = ed or ye
    inc_txns = db.query(Transaction).filter(
        Transaction.type == "income",
        Transaction.date >= _start,
        Transaction.date <= _end,
    ).all()
    exp_txns = db.query(Transaction).filter(
        Transaction.type == "expense",
        Transaction.date >= _start,
        Transaction.date <= _end,
    ).all()
    cash_balance = sum(t.amount for t in inc_txns) - sum(t.amount for t in exp_txns)
    ws_sfp["D65"] = round(max(0.0, cash_balance), 2)

    # 3. Financial liabilities — "Fund from Director" income in the period
    #    (treated as director loans, not equity, per FIRS classification)
    director_loans = sum(
        t.amount for t in inc_txns if t.category == "Fund from Director"
    )
    ws_sfp["D77"] = round(director_loans, 2)

    # 4. Short-term debt — outstanding loan principal not tracked in transactions
    #    (e.g. cooperative society loan balance) — supplied as query parameter
    ws_sfp["D81"] = round(short_term_debt, 2)

    # 5. Equity inputs
    ws_sfp["D112"] = round(share_capital, 2)       # Issued Share Capital
    # D114 (Retained Earnings) links to P&L Retained Earnings C/F (INCOME_STATEMENT!E128)
    ws_sfp["D114"] = "=INCOME_STATEMENT!E128"
    ws_sfp["D124"] = round(other_reserves, 2)       # Other reserves (D119-D124 feed E126)

    # ── Save ──────────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=CIT_Return_{year}.xlsx"},
    )
